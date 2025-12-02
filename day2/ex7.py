import openpyxl

filename = 'video2.xlsx'

wb = openpyxl.load_workbook(filename)
ws = wb['vgsales']

# ws['P1'] = 'Average Sales'
# ws['P2'] = '=AVERAGE(K2:K16328)'
#
# ws['Q1'] = "Number of populated cells"
# ws['Q2'] = "=COUNTA(E2:E16328)"
#
# ws['S1'] = "Total Sports of Sales"
# ws['S2'] = '=COUNTIF(E2:E16328, "sports")'
#
# # print(ws['S2'].value) # =COUNTIF(E2:E16328, "sports"), wypisze formułę
# ws['T1'] = "Total sum of Sports Sales"
# ws['T2'] = '=SUMIF(E2:E16328, "Sports", K2:K16328)'
#
# ws['U1'] = "Rounded sum of Sports Sales"
# ws['U2'] = '=CEILING(T2, 25)'  # zaokrąglenie do najblizzego 25, w górę

# ws['V1'] = "Rounded sum of Sports Sales"
# ws['V2'] = '=CEILING(T2, 1)'  # zaokrąglenie

# ws['W1'] = "Rounded sum of Sports Sales"
# ws['W2'] = '=FLOOR(T2, 25)'  # zaokrąglenie do najblizzego 25, w dół

ws['X1'] = 'Rounded'
ws['X2'] = '=ROUND(T2,0)'

wb.save(filename)
wb.close()
