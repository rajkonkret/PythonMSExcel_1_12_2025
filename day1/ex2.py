# wczytanie pliku excel
from openpyxl import Workbook, load_workbook

workbook = load_workbook('sample.xlsx')
sheet = workbook.active  # wybranie aktywnego arkusza
print(sheet)  # <Worksheet "Sheet">

print(sheet['A1'])  # <Cell 'Sheet'.A1>
print(sheet['A1'].value)
print(sheet['A2'].value)
# 42
# 43

for i in sheet:
    print(i)
    # (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>)
    # (<Cell 'Sheet'.A2>, <Cell 'Sheet'.B2>, <Cell 'Sheet'.C2>)
    # (<Cell 'Sheet'.A3>, <Cell 'Sheet'.B3>, <Cell 'Sheet'.C3>)
    # (<Cell 'Sheet'.A4>, <Cell 'Sheet'.B4>, <Cell 'Sheet'.C4>)

for row in sheet.iter_rows(min_row=1, max_row=4):
    for cell in row:
        print(cell.value)
# 42
# None
# None
# 43
# None
# None
# 1
# 2
# 3
# 4
# 5
# 6

# None - odpowiednik null, brak wartosc
