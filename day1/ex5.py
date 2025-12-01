import openpyxl

filename = "video2.xlsx"

wb = openpyxl.load_workbook(filename)
ws = wb.active

ws = wb['vgsales']

row_position = 1

for i in range(1, ws.max_row):
    row_position += 1
    print(row_position)
    NA_Sales = ws.cell(row=row_position, column=7).value
    EU_Sales = ws.cell(row=row_position, column=8).value
    JP_Sales = ws.cell(row=row_position, column=9).value
    Other_Sales = ws.cell(row=row_position, column=10).value

    total_sales = (NA_Sales + EU_Sales + JP_Sales + Other_Sales)

    ws.cell(row=row_position, column=11).value = total_sales

wb.save(filename)
wb.close()
