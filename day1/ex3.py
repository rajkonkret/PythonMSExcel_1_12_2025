import openpyxl

wb = openpyxl.load_workbook(r"../data/videogamesales.xlsx")
# wb = openpyxl.load_workbook("videogamesales.xlsx")
ws = wb.active

print(wb)  # <openpyxl.workbook.workbook.Workbook object at 0x000001A2248656A0>
print(ws)  # <Worksheet "vgsales">

ws = wb['vgsales']  # ustawienie arkusza do działania

print("Total number of rows:", ws.max_row)  # Total number of rows: 16328
print("Total number of columns:", ws.max_column)  # Total number of columns: 10

print("Values in cell A1 is:", ws['A1'].value)

value = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
print(value)
# ['Rank', 'Name', 'Platform', 'Year', 'Genre', 'Publisher', 'NA_Sales', 'EU_Sales', 'JP_Sales', 'Other_Sales']


wb.close()
