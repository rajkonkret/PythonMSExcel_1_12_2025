# pip - manager pakietów pythona
# pip install openpyxl

from openpyxl import Workbook, load_workbook

wb = Workbook()  # szablon pliku excel
ws = wb.active  # ustawia aktywny arkusz

# komórka A1 w arkuszu
ws['A1'] = 42
ws['A2'] = 43

ws.append([1, 2, 3])
ws.append([4, 5, 6])

# zapisanie danych do pliku excel
wb.save("sample.xlsx")
wb.close()
